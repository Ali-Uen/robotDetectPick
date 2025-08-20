def excel(classOfObject, averageTime, averageDistance, averageSpeed, ox1, oy1, 
          speedofConveyor, timeToPick, x_pick_coordinates, y_pick_coordinates, 
          timeToPerfomAction, robot_x_coordinate_system, robot_y_coordinate_system, 
        world_x_coordinate_system, world_y_coordinate_system, robot_z_coordinate, timeToCalculate):

    home_path = os.path.expanduser('~')
    project_path = os.path.join(home_path, 'robotDetectPick', 'evaluation')
    
    filename = project_path + "/ergebnisse.xlsx"

    def createNewTable():
        from openpyxl import Workbook
        from openpyxl.utils import get_column_letter

        # Erstellen Sie eine neue Arbeitsmappe
        workbook = Workbook()

        # Wählen Sie den aktiven Arbeitsblatt aus
        sheet = workbook.active

        # Definieren Sie die Titel der Spalten
        titel = [
                "Datum",
                "Klasse des Objekts",
                "Durchschnittliche Zeitdifferenz",
                "Zeitperfomance der Pick and Place Aktion",
                "Zeit für das Greifen der Objekte",
                "Zeit für die Geschwindigkeitsberechnung und die linearen Interpolation",
                "Durchschnittliche Distanz",
                "Durchschnittliche Geschwindigkeit cm/s",
                "Geschwindigkeit Förderband (Prozentual)",
                "Erfolgsrate (Ja oder Nein)",
                "Stabilität des gegriffenen Objekts (1-5)",
                "Objekt vertikale Koordinate (Weltkoordinatensystem)",
                "Objekt horizontale Koordinate (Weltkoordinatensystem)",
                "X-Koordinate des Objekts im Roboter Koordinatensystem",
                "Y-Koordinate des Objekts im Roboter Koordinatensystem",
                "x Robot(oben links, gegen Uhrzeigersinn)",
                "y Robot(oben links, gegen Uhrzeigersinn)",
                "x World(oben links, gegen Uhrzeigersinn)",
                "y World(oben links, gegen Uhrzeigersinn)",
                "durchschnittliche Z-Koordinate (Roboter)",
        ]

        # Fügen Sie die Titel in die entsprechenden Zellen ein und stellen Sie die Spaltenbreite ein
        for col, titel_text in enumerate(titel, start=1):
            zelle = sheet.cell(row=1, column=col)
            zelle.value = titel_text
            sheet.column_dimensions[get_column_letter(col)].width = 40

        # Speichern Sie die Arbeitsmappe
        workbook.save(filename)


    from datetime import datetime
    from openpyxl import load_workbook
    from openpyxl.utils import get_column_letter
    import os


    if not os.path.exists(filename):
        createNewTable()

    # Laden Sie die vorhandene Arbeitsmappe
    workbook = load_workbook(filename)

    # Wählen Sie den aktiven Arbeitsblatt aus
    sheet = workbook.active
    # Definieren Sie die Titel der Spalten
    # Fügen Sie das aktuelle Datum und die Uhrzeit in die letzte Spalte ein
    zeitstempel = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    titel = [
        "Datum",
        "Klasse des Objekts",
        "Durchschnittliche Zeitdifferenz",
        "Zeitperfomance der Pick and Place Aktion",
        "Zeit für das Greifen der Objekte",
        "Zeit für die Geschwindigkeitsberechnung und die linearen Interpolation",
        "Durchschnittliche Distanz",
        "Durchschnittliche Geschwindigkeit cm/s",
        "Geschwindigkeit Förderband (Prozentual)",
        "Erfolgsrate (Ja oder Nein)",
        "Stabilität des gegriffenen Objekts (1-5)",
        "Objekt vertikale Koordinate (Weltkoordinatensystem)",
        "Objekt horizontale Koordinate (Weltkoordinatensystem)",
        "X-Koordinate des Objekts im Roboter Koordinatensystem",
        "Y-Koordinate des Objekts im Roboter Koordinatensystem",
        "x Robot(oben links, gegen Uhrzeigersinn)",
        "y Robot(oben links, gegen Uhrzeigersinn)",
        "x World(oben links, gegen Uhrzeigersinn)",
        "y World(oben links, gegen Uhrzeigersinn)",
        "durchschnittliche Z-Koordinate (Roboter)",

    ]
    # Fügen Sie die Titel in die entsprechenden Zellen ein und stellen Sie die Spaltenbreite ein
    for col, titel_text in enumerate(titel, start=1):
        zelle = sheet.cell(row=1, column=col)
        zelle.value = titel_text
        sheet.column_dimensions[get_column_letter(col)].width = 40

    # Verfolgen Sie die Werte der Variablen im Echtzeit-Skript
    variable_werte = {
        "Datum": [zeitstempel],
        "Klasse des Objekts": [classOfObject],
        "Durchschnittliche Zeitdifferenz": [averageTime],
        "Zeitperfomance der Pick and Place Aktion": [timeToPerfomAction],
        "Zeit für das Greifen der Objekte": [timeToPick],
        "Zeit für die Geschwindigkeitsberechnung und die linearen Interpolation": [timeToCalculate],
        "Durchschnittliche Distanz": [averageDistance],
        "Durchschnittliche Geschwindigkeit cm/s": [averageSpeed],
        "Geschwindigkeit Förderband (Prozentual)": [speedofConveyor],
        "Erfolgsrate (Ja oder Nein)": [],
        "Stabilität des gegriffenen Objekts (1-5)": [],
        "Objekt vertikale Koordinate (Weltkoordinatensystem)": [ox1],
        "Objekt horizontale Koordinate (Weltkoordinatensystem)":[oy1],
        "X-Koordinate des Objekts im Roboter Koordinatensystem": [x_pick_coordinates],
        "Y-Koordinate des Objekts im Roboter Koordinatensystem": [y_pick_coordinates],
        "x Robot(oben links, gegen Uhrzeigersinn)": [robot_x_coordinate_system[0]],
        "y Robot(oben links, gegen Uhrzeigersinn)": [robot_y_coordinate_system[0]],
        "x World(oben links, gegen Uhrzeigersinn)": [world_x_coordinate_system[0]],
        "y World(oben links, gegen Uhrzeigersinn)": [world_y_coordinate_system[0]],
        "durchschnittliche Z-Koordinate (Roboter)": [robot_z_coordinate],


    }

    # Ermitteln Sie die letzte Zeile für jede Spalte
    letzte_zeile = {}
    for variable, werte in variable_werte.items():
        spalte = titel.index(variable) + 1
        letzte_zeile[spalte] = sheet.max_row + 1

    # Fügen Sie die Variablenwerte in die entsprechenden Zeilen und Spalten ein
    for variable, werte in variable_werte.items():
        spalte = titel.index(variable) + 1
        start_zeile = letzte_zeile.get(spalte, 2)  # Beginnen Sie mit Zeile 2, falls keine vorherigen Werte in der Spalte vorhanden sind

        for wert in werte:
            if isinstance(wert, list):
                # Fügen Sie die Elemente der Arrays zeilenweise ein
                for element in wert:
                    zelle = sheet.cell(row=start_zeile, column=spalte)
                    zelle.value = element

                    start_zeile += 1
                    letzte_zeile[spalte] = start_zeile
            else:
                zelle = sheet.cell(row=start_zeile, column=spalte)
                zelle.value = wert

                start_zeile += 1
                letzte_zeile[spalte] = start_zeile




    # Speichern Sie die aktualisierte Arbeitsmappe
    workbook.save(filename)




# You can use this evaluation function to create always new tables when detect an object

def excelWithDifferentNewTables(averageTime, averageSpeed, ox1, speedofConveyor, timeToPick, x_pick_coordinates, y_pick_coordinates, timeToPerfomAction):
    import os
    from openpyxl import Workbook, load_workbook
    from openpyxl.utils import get_column_letter

    # Definieren Sie den ursprünglichen Dateinamen
    file_name = "/home/niryo/ali/bachelorarbeit/evaluation/ergebnisse"

    # Funktion zur Überprüfung und Generierung eines eindeutigen Dateinamens
    def get_unique_filename(base_name):
        """Erzeugt einen eindeutigen Dateinamen, indem eine chronologische Nummer angehängt wird."""
        if not os.path.exists(base_name + ".xlsx"):
            return base_name

        suffix = 2
        while True:
            unique_name = f"{base_name}{suffix}.xlsx"
            if not os.path.exists(unique_name):
                return unique_name
            suffix += 1

    # Generieren Sie einen eindeutigen Dateinamen
    unique_file_name = get_unique_filename(file_name)

    # Laden Sie die vorhandene Arbeitsmappe oder erstellen Sie eine neue
    if os.path.exists(unique_file_name):
        workbook = load_workbook(unique_file_name)
    else:
        workbook = Workbook()

    # Wählen Sie den aktiven Arbeitsblatt aus
    sheet = workbook.active

    # Definieren Sie die Titel der Spalten
    titel = [
        "Durchschnittliche Zeitdifferenz",
        "Durchschnittliche Geschwindigkeit cm/s",
        "Objekt vertikale Koordinate (Weltkoordinatensystem)",
        "Erfolgsrate (Ja oder Nein)",
        "Geschwindigkeit Förderband (Prozentual)",
        "Stabilität des gegriffenen Objekts (1-5)",
        "Zeit für das Greifen der Objekte",
        "X-Koordinate des Objekts im Roboter Koordinatensystem",
        "Y-Koordinate des Objekts im Roboter Koordinatensystem",
        "Zeitperfomance der Pick and Place Aktion"
    ]

    # Fügen Sie die Titel in die entsprechenden Zellen ein und stellen Sie die Spaltenbreite ein
    for col, titel_text in enumerate(titel, start=1):
        zelle = sheet.cell(row=1, column=col)
        zelle.value = titel_text
        sheet.column_dimensions[get_column_letter(col)].width = 40

    # Speichern Sie die Arbeitsmappe
    workbook.save(unique_file_name)

    # Verfolgen Sie die Werte der Variablen im Echtzeit-Skript
    variable_werte = {
        "Durchschnittliche Zeitdifferenz": [averageTime],
        "Durchschnittliche Geschwindigkeit cm/s": [averageSpeed],
        "Objekt vertikale Koordinate (Weltkoordinatensystem)": [ox1],
        "Erfolgsrate (Ja oder Nein)": [],
        "Geschwindigkeit Förderband (Prozentual)": [speedofConveyor],
        "Stabilität des gegriffenen Objekts (1-5)": [],
        "Zeit für das Greifen der Objekte": [timeToPick],
        "X-Koordinate des Objekts im Roboter Koordinatensystem": [x_pick_coordinates],
        "Y-Koordinate des Objekts im Roboter Koordinatensystem": [y_pick_coordinates],
        "Zeitperfomance der Pick and Place Aktion": [timeToPerfomAction]
    }

    # Laden Sie die aktualisierte Arbeitsmappe erneut
    workbook = load_workbook(unique_file_name)

    # Wählen Sie den aktiven Arbeitsblatt aus
    sheet = workbook.active

    # Ermitteln Sie die letzte Zeile für jede Spalte
    letzte_zeile = {}
    for variable, werte in variable_werte.items():
        spalte = titel.index(variable) + 1
        letzte_zeile[spalte] = sheet.max_row + 1

    # Fügen Sie die Variablenwerte in die entsprechenden Zeilen und Spalten ein
    for variable, werte in variable_werte.items():
        spalte = titel.index(variable) + 1
        start_zeile = letzte_zeile.get(spalte, 2)  # Beginnen Sie mit Zeile 2, falls keine vorherigen Werte in der Spalte vorhanden sind

        for wert in werte:
            zelle = sheet.cell(row=start_zeile, column=spalte)
            zelle.value = wert

            start_zeile += 1
            letzte_zeile[spalte] = start_zeile

    # Speichern Sie die aktualisierte Arbeitsmappe
    workbook.save(unique_file_name)